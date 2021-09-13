import csv
import gc
import os
import shutil as shutilb
import sys
from datetime import datetime, timedelta

import folium
import folium.plugins as plugins
import geopandas as gpd
import numpy as np
import pandas as pd
import rasterio
import shapely
import spotpy
from branca.element import Template, MacroElement
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles import Font
from pydsstools._lib.x64.py37.core_heclib import TimeSeriesContainer, squeeze_file
from pydsstools.heclib.dss import HecDss
from rasterio import features, shutil, mask
from rasterstats import zonal_stats
from shapely import affinity

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ USER-DEFINED VARIABLES AND INPUT ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# Variable 'base_dir' is the project directory where this script resides, needs editing by user
base_dir = r'C:\Projects\design-storm'

# Variable 'hms_model_dir' is a directory for the HMS model, needs editing by user
hms_model_dir = base_dir + os.sep + r'HmsProject\Trin_HMS'

# Variable 'hms_project' is the project name of the HMS model, needs editing by user
hms_project = "CWMS_Trinity_River"

# Variable 'hms_run' is the name of the HMS simulation created for design storm optimization
hms_run = "DesignStorm"

# Variable 'ctrl_spec_time_interval' is the HEC-HMS Control Specifications Time Interval, i.e. use '15MIN' or '1HOUR'
ctrl_spec_time_interval = '1HOUR'

# Variable 'hms_met_model' is the name of the HMS meteorological model (specified hyetograph) for design storm runs
hms_met_model = "DesignStorm.met"

# Variable 'start_time_str' is a string for the start time of HMS simulation, the user needs to edit and keep a format
# of 'ddmmmyyyy,HHMMSS'.
start_time_str = "01Jan2000,000000"

# 'dss_in' is the path and name of the input HEC-HMS DSS file of subbasin specified hyetographs
dss_in = hms_model_dir + os.sep + r'data\hyetographs.dss'

# 'dss_out' is the path and name of the output HEC-HMS DSS file with computation results
# The output DSS file by default is named after the HMS run, with hyphens and spaces replaced with underscores
dss_out = hms_model_dir + os.sep + hms_run.replace("-", "_").replace(" ", "_") + '.dss'

# Variable 'junction_list' is a tuple of junction names (strings) that need to be exactly those in the HMS models
# The user needs to prepare the junction names of interest; junction_list must contain at least 2 junctions.
junction_list = ('Trinity_River_J010', 'Trinity_River_at_Trinidad')

# Variable 'junction_shapefile' is the path and name of the junction shapefile, exported from HEC-HMS, defined by user
junction_shapefile = base_dir + os.sep + r'GisData\BasinShapefiles\Trinity_CWMS_Junctions.shp'

# Variable 'junction_col_name' represents the name of the column in 'junction_shapefile' specified above where the
# junction names are stored. This needs editing by the user to exactly match the shapefile column name.
junction_col_name = 'Name'

# Variable 'runs' is the limit of optimization runs for each junction. Along with HMS simulation time,
# it determines the total run time. Testing on a 15,000 sqmi basin indicates 300 is a good number, which
# gave a total run time of ~ 1 hour per junction.
runs = 300

# Variable 'frequency' is the return period of the analysis, defined by user
frequency = '100yr'  # update for each frequency run!!!!!!

# Variable 'duration' is the total storm duration of the analysis, defined by user
duration = '48hr'

# Variable 'basin_name' is the name of the study basin, used to label output, defined by user
basin_name = "Trinity"

# Variable 'output_name' is a unique identifier used to name this script's output files, defined by user
# 'output_name' should change with each script 3 run, otherwise previously computed results may be overwritten
output_name = frequency + duration + "_flow"

# Variable 'output_dir' is the directory where output from this script will be located
output_dir = base_dir + os.sep + r'ScriptResults\3_Run_HMS_Elliptical_Design_Storm_Optimization'

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ USER-DEFINED INPUT NEEDED FROM SCRIPTS 1 & 2 ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# Variable 'ellipses_shapefile' is the path and file name of the DAR shapefile (output from script 1), defined by user
ellipses_shapefile = base_dir + os.sep + r'ScriptResults\1_Prepare_DAR_Ellipses\1a_Ellipses_Vector_dar48hr.shp'

# Variable 'na14_raster' is the path and name of the precipitation frequency .asc file, defined by user
# This is the total duration .asc file that was projected in script 2.
na14_raster = base_dir + os.sep + r'GisData\PrecipNA14\tx100yr48ha_ams_proj.asc'

# Variable 'subbasin_precip_shp' is the path and name of the subbasin shapefile (output from script 2), defined by user
subbasin_precip_shp = base_dir + os.sep + \
                      r'ScriptResults\2_Prepare_NA14_Precipitation\2a_Subbasins_Precip_Vector_100yr.shp'

# Variable 'subbasin_precip_csv' is the path and name of the subbasin csv (output from script 2), defined by user
subbasin_precip_csv = base_dir + os.sep + \
                      r'ScriptResults\2_Prepare_NA14_Precipitation\2a_Subbasins_Precip_Vector_100yr.csv'

# Variable 'subbasin_col_name' represents the name of the column in the 'subbasin_precip_shp' specified above where the
# subbasin names are stored. This needs editing by the user to exactly match the shapefile column name.
subbasin_col_name = 'Name'

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ INITIALIZATION BASED ON PRIOR USER INPUT ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# This section, and beyond, includes variables that should generally not be changed. Thus no special editing is needed.

# Variable 'hyetograph_time_step' is the interval of the hyetograph in hours.
hyetograph_time_step = 1  # script currently setup to run at 1 hour hyetograph time step only

# Variable 'theta0' is the initial orientation of the design storm, '0' means the storm is positioned horizontally.
# Degrees are measured counter clockwise from the positive x-axis
theta0 = 0

# Variable 'software_dir' is a directory for various software
software_dir = base_dir + os.sep + 'Software'

# Variable 'hms_install_dir' is the installation directory for HEC-HMS.exe
hms_install_dir = software_dir + os.sep + 'HEC-HMS-4.8'

# Variable 'jdk_install_dir' is the installation directory for Java (jdk)
# The jdk is needed to call HMS headlessly
jdk_install_dir = software_dir + os.sep + 'jdk-11.0.10'

# Variable 'jython_install_exe' is the installation directory for Jython
# Jython is needed to call HMS headlessly
jython_install_exe = software_dir + os.sep + r'jython-2.7.2\bin\jython.exe'

# 'temp_dir' is a directory where temporary files are stored
temp_dir = software_dir + os.sep + 'TempFiles'

# 'csv_out_2' is the name of this script's output file with optimized storm locations and peak flows for each junction
csv_out_2 = temp_dir + os.sep + '3b_' + basin_name + '_XYThetaOpt_' + output_name + '.csv'

# Remove previous output file if it exists
if os.path.exists(csv_out_2):
    os.remove(csv_out_2)

# Read in subbasin files from script 2 and use the centroid coordinate to set the initial storm center
df_basin = pd.read_csv(subbasin_precip_csv, converters={'Temporal': eval})
gdf_basin = gpd.read_file(subbasin_precip_shp)
gdf_basin['Temporal'] = df_basin['Temporal']

gdf_basin['Precip'] = 0
basin_dissolved = gdf_basin.copy()
basin_dissolved = basin_dissolved.dissolve(by='Precip')
basin_centroid = basin_dissolved['geometry'].centroid

x0 = float(basin_centroid.x)
y0 = float(basin_centroid.y)

# Get the bounds of the basin outline to set the allowable area for optimization
bounds = basin_dissolved.total_bounds

x_low = bounds[0]
x_high = bounds[2]
y_low = bounds[1]
y_high = bounds[3]

del basin_dissolved

# 'jython_script' is the name of script file used to run HMS automatically.
jython_script = temp_dir + os.sep + 'HMScompute.py'

# 'start_time_num' is the numerical form of the simulation start time.
start_time_num = datetime.strptime(start_time_str, '%d%b%Y,%H%M%S')

# 'hms_metmodel_backup' is a copy of the hms met model. The met model is copied over before each run to avoid a bug that
# sometimes deletes precip gages from the met model file when hms is called repeatedly
shutilb.copy(hms_model_dir + os.sep + hms_met_model, temp_dir)
hms_metmodel_backup = temp_dir + os.sep + hms_met_model

# ~~~~~~~~~~~~~~~~~~~ PREPARATION: WRITE AN HMSCOMPUTE.PY JYTHON SCRIPT FILE ~~~~~~~~~~~~~~~~~~~

# Write a Jython script file that simulates the prepared simulation in the target HMS model
file = open(jython_script, "w")
file.write("from hms.model import Project" + "\n")
file.write("from hms import Hms" + "\n")
file.write("myProject = Project.open('" + hms_model_dir + os.sep + hms_project + ".hms')" + "\n")
file.write("myProject.computeRun('" + hms_run + "')" + "\n")
file.write("myProject.close()" + "\n")
file.write("Hms.shutdownEngine()" + "\n")
file.close()
print("\n***** Preparing HMScompute.py jython script for HMS simulation run... *****\n")

# ~~~~~~~~~~~~~~~~~~~~~PREPARATION: WRITE AN HMSCOMPUTE.BAT FILE~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# Write an HMScompute.bat file that calls HMS and executes the HMScompute.py Jython file
file = open(temp_dir + os.sep + "HMScompute.bat", "w")
file.write("@echo off" + "\n")
file.write('\n')
file.write('set "HMS_HOME=' + hms_install_dir + '"' + '\n')
file.write(r'set "PATH=%HMS_HOME%\bin\gdal;%PATH%"' + '\n')
file.write(r'set "GDAL_DRIVER_PATH=%HMS_HOME%\bin\gdal\gdalplugins"' + '\n')
file.write(r'set "GDAL_DATA=%HMS_HOME%\bin\gdal\gdal-data"' + '\n')
file.write(r'set "PROJ_LIB=%HMS_HOME%\bin\gdal\projlib"' + '\n')
file.write(r'set "CLASSPATH=%HMS_HOME%\hms.jar;%HMS_HOME%\lib\*"' + '\n')
file.write('set "JAVA_HOME=' + jdk_install_dir + '"' + '\n')
file.write('\n')
file.write(r'%HMS_HOME%\jre\bin\java -Djava.library.path=%HMS_HOME%\bin;%HMS_HOME%\bin\gdal;%HMS_HOME%\bin\hdf; '
           r'org.python.util.jython ' + jython_script + '\n')
file.close()
print("\n***** Preparing .bat file that executes HMScompute.py jython script... *****\n")

# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ DEFINE THE OPTIMIZATION FUNCTION ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# In this section, the cost function is defined for the optimization process
# The optimization process aims to minimize the cost function
# The script calls an external library called spotpy and uses existing functions within
class SpotpySetup(object):
    # Define the initialization function
    def __init__(self):
        # Introduce some global variables
        global x_high, x_low, y_high, y_low
        # Set the boundaries, best guesses and distribution of the three parameters, X, Y, and theta
        self.params = [spotpy.parameter.Uniform('X', low=x_low, high=x_high, optguess=(x_low + x_high) / 2.0),
                       spotpy.parameter.Uniform('Y', low=y_low, high=y_high, optguess=(y_low + y_high) / 2.0),
                       spotpy.parameter.Uniform('theta', low=0, high=180, optguess=0)]

    # Set the mandatory parameters function
    def parameters(self):
        return spotpy.parameter.generate(self.params)

    # Set the mandatory simulation function, which includes 4 steps: create storm, prepare HMS input DSS, run HMS,
    # and extract peak flow from the output DSS.
    def simulation(self, x):
        # Introduce some global variables
        global base_dir, ellipses_shapefile, x0, y0, theta0, hyetograph_time_step, output_dir, \
            software_dir, dss_in, junction, subbasin_col_name, hms_model_dir, \
            na14_raster, start_time_num, dss_out

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ CREATE TOTAL STORM ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        # The DAR ellipses are rotated and transposed to a new location, then rasterized
        # Then the underlying NA14 precipitation point frequency depths are queried and multiplied by the DAR raster
        # Then the average, total storm precipitation for each subbasin is calculated

        print("\n***** Rotating and Translating Storm... *****\n")
        # Read in the ellipses shapefile from script 1
        gdf_ellipse = gpd.read_file(ellipses_shapefile)
        gdf_ellipse.reset_index(drop=True)

        # Rotate the ellipses
        ellipse_rotated = gdf_ellipse.copy()
        for index, row in ellipse_rotated.iterrows():
            rotated = shapely.affinity.rotate(row['geometry'], angle=theta0 - x[2])
            ellipse_rotated.loc[index, 'geometry'] = rotated

        # Translate the ellipses
        ellipse_translated = ellipse_rotated.copy()
        for index, row in ellipse_translated.iterrows():
            translated = shapely.affinity.translate(row['geometry'], xoff=x[0] - x0, yoff=x[1] - y0)
            ellipse_translated.loc[index, 'geometry'] = translated

        # Mask the reprojected na14 precipitation raster based on rotated/translated ellipses
        with rasterio.open(na14_raster) as template:
            out_precip, out_transform = rasterio.mask.mask(dataset=template, shapes=ellipse_translated.geometry,
                                                           crop=True)
            no_data = template.nodata
            out_precip[out_precip == no_data] = 0
            precip_meta = template.meta.copy()
            precip_meta.update({"driver": "AAIGrid",
                                "height": out_precip.shape[1],
                                "width": out_precip.shape[2],
                                "transform": out_transform,
                                "nodata": 0,
                                "dtype": "int32"})

        # Filename of the clipped na14 raster that will be created
        na14_clipped = temp_dir + os.sep + 'na14_clip.asc'

        # If it exists, delete a previously masked na14 raster file
        if os.path.isfile(na14_clipped):
            try:
                rasterio.shutil.delete(na14_clipped)
            except BaseException as error:
                print('An exception occurred: {}'.format(error))

        with rasterio.open(na14_clipped, "w", **precip_meta) as destination:
            destination.write(out_precip)

        # Filename of the transposed dar raster that will be created
        dar_transposed = temp_dir + os.sep + 'dar_transposed.tif'

        # Burn the rotated/translated ellipse features into a raster
        dar_meta = precip_meta.copy()
        dar_meta.update({"driver": "GTiff",
                        "nodata": 0,
                        "dtype": "float64"})

        with rasterio.open(dar_transposed, 'w+', **dar_meta) as dar_transposed_out:
            out_array = dar_transposed_out.read(1)
            shapes = ((geom, value) for geom, value in zip(ellipse_translated.geometry, ellipse_translated.dar))
            burned = features.rasterize(shapes=shapes, fill=0, out=out_array, transform=dar_transposed_out.transform)
            dar_transposed_out.write_band(1, burned)

        # Multiply the masked na14 precip by the transposed dar raster, divide by 1000 to convert to decimal inches
        with rasterio.open(dar_transposed) as source:
            dar = source.read(1, masked=True)

        with rasterio.open(na14_clipped) as source:
            masked_precip = source.read(1, masked=True)

        reduced_precip = masked_precip * dar / 1000.00
        reduced_precip_meta = dar_meta.copy()

        # Filename of the reduced storm raster that will be created
        storm_final = temp_dir + os.sep + 'Storm_' + junction + '.tif'

        with rasterio.open(storm_final, 'w+', **reduced_precip_meta) as outRaster2:
            outRaster2.write_band(1, reduced_precip)

        # Calculate zonal statistics to determine the average precipitation for each subbasin
        print("\n***** Calculating zonal statistics for the total, average precipitation at each subbasin... *****\n")
        with rasterio.open(storm_final, 'r+') as source:
            array = source.read(1)
            affine = source.transform
        stats = zonal_stats(gdf_basin, array, affine=affine, nodata=-9999, stats='mean', band=1)
        mean_values = []
        for entry in stats:
            mean_value = entry.get('mean')
            mean_values.append(mean_value)

        # Delete precipitation rasters created during this iteration
        try:
            rasterio.shutil.delete(na14_clipped)
            rasterio.shutil.delete(dar_transposed)
            rasterio.shutil.delete(storm_final)
        except BaseException as error:
            print('An exception occurred: {}'.format(error))

        # Add subbasin average precip values to geo dataframe
        gdf_basin['Precip'] = mean_values
        gdf_basin['Precip'] = gdf_basin['Precip'].fillna(0)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ PREPARE HMS HYETOGRAPH INPUT ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        # Write specified hyetograph data to dss
        start_date_time = (start_time_num + timedelta(hours=1)).strftime("%d%b%Y %H:%M:%S").upper()
        temporal_length = len(gdf_basin['Temporal'][0])
        for index, row in gdf_basin.iterrows():
            depths = []
            for duration in range(temporal_length):
                depths.append(row['Precip'] * row['Temporal'][duration] / sum(row['Temporal']))
            # Use pydsstools library to convert Pandas dataframe to DSS
            pathname = "//" + row[subbasin_col_name].upper() + "/PRECIP-INC//" + str(hyetograph_time_step).upper() \
                       + "HOUR//"
            tsc = TimeSeriesContainer()
            tsc.pathname = pathname
            tsc.startDateTime = start_date_time
            tsc.numberValues = temporal_length
            tsc.units = "IN"
            tsc.type = "PER-CUM"
            tsc.interval = hyetograph_time_step  # hours
            tsc.values = depths

            fid = HecDss.Open(dss_in)
            fid.deletePathname(tsc.pathname)
            fid.put_ts(tsc)
            fid.close()
        squeeze_file(dss_in)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ SIMULATE HMS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        print("\n***** Created the specified hyetographs DSS file; Calling HEC-HMS... *****\n")
        os.chdir(temp_dir)
        # Copy over the HMS met model file to avoid a bug that sometimes removes precip gages from the met model
        os.remove(hms_model_dir + os.sep + hms_met_model)
        shutilb.copy(hms_metmodel_backup, hms_model_dir)
        sys_command = ".\\HMScompute.bat"
        os.system(sys_command)

        # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ EXTRACT PEAK FLOW ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

        # Extract the junction peak flow from the DSS file holding model output
        print("\n***** Reading the output DSS file... *****\n")
        pathname = "//" + junction.upper() + "/FLOW//" + ctrl_spec_time_interval.upper() + "/RUN:" + hms_run + "/"

        fid = HecDss.Open(dss_out)
        time_series = fid.read_ts(pathname)
        values = time_series.values
        try:
            max_val = max(values)
        except TypeError:
            print("The specified 'dss_out' file and 'pathname' cannot be found.")
            sys.exit(1)
        fid.close()
        squeeze_file(dss_out)

        # Save junction peak flow value as csv
        row = [round(max_val), round(x[0], 2), round(x[1], 2), round(x[2])]
        with open(csv_out, 'a') as csv2:
            writer2 = csv.writer(csv2, lineterminator='\n')
            writer2.writerow(row)

        # Check HEC-HMS run log for errors; write error to script error log if found
        hms_run_log = hms_model_dir + os.sep + os.path.splitext(os.path.basename(dss_out))[0] + ".log"
        script_error_log = output_dir + os.sep + "0_HEC-HMS_ERROR_LOG_" + hms_run + ".txt"
        if os.path.exists(hms_run_log):
            with open(hms_run_log, 'r') as run_log:
                for line in run_log:
                    if line.startswith("ERROR"):
                        if os.path.exists(script_error_log):
                            write_mode = 'a'
                        else:
                            write_mode = 'w'
                        with open(script_error_log, write_mode) as error_log:
                            error_log.write(line)
                            error_log.write("The below simulation results for " + junction + " are invalid: \n")
                            error_log.write(str(row) + "\n")
                            error_log.write("\n")
                        print("\n***** HEC-HMS ERROR!!!!! Check error log in output directory... *****\n")

        # The return value of the cost function is defined as below, note that the cost function is minimized
        # so that the peak flow is maximized
        simulations = [(1 / max_val) * 10000]

        return simulations

    # Define a dummy function 'evaluation' which simply returns 0
    def evaluation(self):
        observations = [0]
        return observations

    # Define a function 'objectivefunction', which calcuates the RMSE between the return values of the 'evaluation'
    # and the 'simulation' functions.
    def objectivefunction(self, simulation, evaluation):
        objectivefunction = spotpy.objectivefunctions.rmse(evaluation, simulation)
        return objectivefunction


# ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ OPTIMIZATION PROCESS ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

# This section is the main part of the script that calls the previously defined functions and optimizes.

# Create a file that will store optimized parameters and flows for each junction
with open(csv_out_2, 'w') as csv_file_2:
    q_writer = csv.writer(csv_file_2, lineterminator='\n')
    q_writer.writerow(['Junction', 'PeakFlow', 'X', 'Y', 'Theta'])

# Go through all the junctions
counter = 0
for i in range(len(junction_list)):
    # Assign the current junction name
    junction = junction_list[i]
    # Prepare a csv file to log peak flow, lon, lat, theta for ALL runs (to be used to create an optimization map)
    csv_out = temp_dir + os.sep + "3a_" + junction + "_RunLog_" + output_name + ".csv"
    with open(csv_out, 'w') as csvFile:
        writer = csv.writer(csvFile, lineterminator='\n')
        writer.writerow(['PeakFlow', 'X', 'Y', 'Theta'])

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ OPTIMIZE FOR THE CURRENT JUNCTION ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    sampler = spotpy.algorithms.sceua(SpotpySetup(), dbname=temp_dir + os.sep + '3_Spotpy_DB_SCEUA_' + junction,
                                      dbformat='csv', db_precision=np.float32)
    sampler.sample(runs, ngs=4)
    collected = gc.collect()

    if os.path.exists(temp_dir + os.sep + '3_Spotpy_DB_SCEUA_' + junction + '.csv'):
        os.remove(temp_dir + os.sep + '3_Spotpy_DB_SCEUA_' + junction + '.csv')

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ WRITE AN OUTPUT CSV FILE w/ OPTIMIZED INFO~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    df_runlog = pd.read_csv(csv_out, dtype={'PeakFlow': np.float32, 'X': np.float32,
                                            'Y': np.float32, 'Theta': np.float32})
    max_flow_index = df_runlog['PeakFlow'].idxmax()
    max_flow_row = df_runlog.loc[max_flow_index, :]
    with open(csv_out_2, 'a+') as csv_file_2:
        q_writer = csv.writer(csv_file_2, lineterminator='\n')
        q_writer.writerow([junction_list[i], round(max_flow_row[0]), round(max_flow_row[1], 2),
                           round(max_flow_row[2], 2), round(max_flow_row[3])])

    # Create a geo-dataframe with all of the simulation points, save it as a shapefile
    gdf_runlog = gpd.GeoDataFrame(df_runlog,
                                  geometry=gpd.points_from_xy(x=df_runlog['X'], y=df_runlog['Y'], crs=gdf_basin.crs))
    gdf_runlog.to_file(output_dir + os.sep + '3c_Optimization_Points_' + junction + '_' + output_name + '.shp')

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ CREATE AN OPTIMIZATION MAP ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~

    print("\n***** Creating a map... *****\n")

    # Calculate percentiles for coloring
    gdf_runlog = gdf_runlog.to_crs(epsg='4326')
    p0 = gdf_runlog['PeakFlow'].min().min()
    p75 = gdf_runlog['PeakFlow'].quantile(0.75)
    p95 = gdf_runlog['PeakFlow'].quantile(0.95)
    p100 = gdf_runlog['PeakFlow'].max().max()

    colors = []

    try:
        for index, row in gdf_runlog.iterrows():
            current = row['PeakFlow']
            if current <= p75:
                colors.append('green')
            elif current <= p95:
                colors.append('yellow')
            elif current < p100:
                colors.append('red')
            elif current == p100:
                colors.append('purple')
            else:
                colors.append('white')
    except:
        colors.append('white')

    gdf_runlog['Color'] = colors

    # Simplify basin geometry for file size efficiency
    gdf_basin_map = gdf_basin.copy()
    gdf_basin_map = gdf_basin_map[[subbasin_col_name, 'geometry']]
    gdf_basin_map['geometry'] = gdf_basin_map['geometry'].simplify(120)

    # Identify center coordinates for map
    gdf_basin_map = gdf_basin_map.to_crs(epsg='4326')
    bounds_lat_lon = gdf_basin_map.total_bounds
    lon_low = bounds_lat_lon[2]
    lon_high = bounds_lat_lon[0]
    lat_low = bounds_lat_lon[1]
    lat_high = bounds_lat_lon[3]

    map_center_lon = (lon_low + lon_high) / 2.0
    map_center_lat = (lat_low + lat_high) / 2.0

    # Create a folium map object
    mapa = folium.Map([map_center_lat, map_center_lon],
                      zoom_start=8,
                      tiles='cartodbpositron',
                      control_scale=True
                      )

    # Add basin layer to map
    style = {'color': 'gray', 'fillColor': 'gray', 'fillOpacity': 0.2, 'weight': 1.0}
    folium.GeoJson(data=gdf_basin_map,
                   name=basin_name + " Basin",
                   popup=gdf_basin_map[subbasin_col_name],
                   style_function=lambda x: style,
                   ).add_to(mapa)

    # Add hover functionality to basin layer
    style_function = lambda x: {'fillColor': '#ffffff',
                                'color': '#000000',
                                'fillOpacity': 0.001,
                                'weight': 0.1}
    highlight_function = lambda x: {'fillColor': '#000000',
                                    'color': '#000000',
                                    'fillOpacity': 0.50,
                                    'weight': 0.1}
    hover = folium.features.GeoJson(
        data=gdf_basin_map,
        name="Subbasin Overlay",
        style_function=style_function,
        control=True,
        highlight_function=highlight_function,
        tooltip=folium.features.GeoJsonTooltip(
            fields=[subbasin_col_name],
            aliases=['Subbasin Name:'],
            style="background-color: white; color: #333333; font-family: arial; font-size: 12px; padding: 10px;"
        )
    )
    mapa.add_child(hover)
    del gdf_basin_map

    # Add junction layer to map
    gdf_junction = gpd.read_file(junction_shapefile)
    gdf_junction = gdf_junction.to_crs(epsg='4326')
    junction_index = None
    for index, row in gdf_junction.iterrows():
        if row[junction_col_name] == junction:
            junction_index = index
            break

    junction_lon = gdf_junction.loc[junction_index, 'geometry'].centroid.x
    junction_lat = gdf_junction.loc[junction_index, 'geometry'].centroid.y

    feature_group = folium.FeatureGroup(name='Junction: ' + junction)
    folium.Marker(
        location=[junction_lat, junction_lon],
        popup=frequency + duration + ' Peakflow: ' + f'{p100:,}' + 'cfs',
        tooltip=junction,
        icon=folium.Icon(icon="cloud", color="blue")
    ).add_to(feature_group)
    feature_group.add_to(mapa)

    # Add storm center optimization points layer to map
    feature_group2 = folium.FeatureGroup(name='Simulated Storm Centers: ' + junction)
    gdf_runlog['Lat'] = gdf_runlog['geometry'].y
    gdf_runlog['Lon'] = gdf_runlog['geometry'].x
    for index, row in gdf_runlog.iterrows():
        # Create custom popup box
        simulation_lat = str(round(row['Lat'], 3))
        simulation_lon = str(round(row['Lon'], 3))
        simulation_theta = str(round(row['Theta']))
        simulation_flow = f"{(round(row['PeakFlow'])):,}"
        html = f"""
        <h1 style="font-size:20px"> Storm Center Parameters</h1>
        <p> Lat: {simulation_lat}<br>
            Lon: {simulation_lon}<br>
            Angle: {simulation_theta}degrees ccw<br><br>
            This simulated storm center led to a peak flow of <strong>{simulation_flow}cfs</strong> at {junction}.
        <p>
        """.format(simulation_lat=simulation_lat, simulation_lon=simulation_lon, simulation_theta=simulation_theta,
                   simulation_flow=simulation_flow, junction=junction)

        iframe = folium.IFrame(html=html, width=200, height=240)
        popup = folium.Popup(iframe, max_width=360)

        folium.CircleMarker(location=(row['geometry'].y, row['geometry'].x),
                            radius=6,
                            tooltip='Click for simulation results...',
                            popup=popup,
                            color=row['Color'],
                            fill_color=row['Color']
                            ).add_to(feature_group2)
    feature_group2.add_to(mapa)

    mapa.keep_in_front(hover, feature_group, feature_group2)

    # Add heatmap to map
    data_heatmap = gdf_runlog[['Lat', 'Lon', 'PeakFlow']].values.tolist()
    plugins.HeatMap(data=data_heatmap, name='Heat Map', show=False, gradient={0.75: 'lime',
                                                                              0.95: 'yellow',
                                                                              1: 'red'}).add_to(mapa)

    # Add various background tile layers
    folium.TileLayer('cartodbdark_matter', name="cartodb dark", control=True).add_to(mapa)
    folium.TileLayer('openstreetmap', name="open street map", control=True, opacity=0.4).add_to(mapa)
    folium.TileLayer('stamenterrain', name="stamen terrain", control=True, opacity=0.6).add_to(mapa)
    folium.TileLayer('stamenwatercolor', name="stamen watercolor", control=True, opacity=0.6).add_to(mapa)
    folium.TileLayer('stamentoner', name="stamen toner", control=True, opacity=0.6).add_to(mapa)

    # Add a layer controller
    folium.LayerControl(collapsed=True).add_to(mapa)

    # Create a legend using HTML and JavaScript
    template = """
    {% macro html(this, kwargs) %}

    <!doctype html>
    <html lang="en">
    <body>


    <div id='maplegend' class='maplegend' 
        style='position: absolute; z-index:9999; border:2px solid grey; background-color:rgba(255, 255, 255, 0.8);
         border-radius:6px; padding: 10px; font-size:18px; left: 6px; bottom: 40px;'>

    <div class='legend-title'>Legend For Simulated Storm Centers</div>
    <div class='legend-scale'>
      <ul class='legend-labels'>
        <li><span style='background:purple;opacity:0.7;'></span>Optimized Peak Flow At Junction</li>
        <li><span style='background:red;opacity:0.7;'></span>Very High Peak Flow At Junction</li>
        <li><span style='background:yellow;opacity:0.7;'></span>High Peak Flow At Junction</li>
        <li><span style='background:green;opacity:0.7;'></span>Low Peak Flow At Junction</li>
        <li><span style='background:gray;opacity:0.5;'></span>Basin</li>


      </ul>
    </div>
    </div>

    </body>
    </html>

    <style type='text/css'>
      .maplegend .legend-title {
        text-align: left;
        margin-bottom: 5px;
        font-weight: bold;
        font-size: 90%;
        }
      .maplegend .legend-scale ul {
        margin: 0;
        margin-bottom: 5px;
        padding: 0;
        float: left;
        list-style: none;
        }
      .maplegend .legend-scale ul li {
        font-size: 80%;
        list-style: none;
        margin-left: 0;
        line-height: 18px;
        margin-bottom: 2px;
        }
      .maplegend ul.legend-labels li span {
        display: block;
        float: left;
        height: 16px;
        width: 30px;
        margin-right: 5px;
        margin-left: 0;
        border: 1px solid #999;
        }
      .maplegend .legend-source {
        font-size: 80%;
        color: #777;
        clear: both;
        }
      .maplegend a {
        color: #777;
        }
    </style>
    {% endmacro %}"""

    macro = MacroElement()
    macro._template = Template(template)
    mapa.get_root().add_child(macro)

    # Save map to html
    mapa.save(output_dir + os.sep + '3d_Optimization_Map_' + junction + '_' + output_name + '.html')

    # ~~~~~~~~~~~~~~~~~~~~~~~~~~~~~ CREATE COLORRAMPS IN AN EXCEL WORKBOOK ~~~~~~~~~~~~~~~~~~~~~~~~~~

    # Combine log .csv files into one Excel workbook; add color gradient; bold font the highest peak flow

    print("\n***** Combining simulation log files into a color ramp .xlsx file... *****\n")
    try:
        from openpyxl.cell import get_column_letter
    except ImportError:
        from openpyxl.utils import get_column_letter
    xlsx_out = output_dir + os.sep + "3a_" + basin_name + "_RunLogAll_" + output_name + ".xlsx"
    if counter == 0:
        writer = pd.ExcelWriter(xlsx_out)
        pd.read_csv(csv_out).to_excel(writer, sheet_name=junction, index=False)
        writer.save()
        writer.close()
        wb = load_workbook(filename=xlsx_out)
        ws = wb[junction]
        ws.conditional_formatting.add('A1:A1000',
                                      ColorScaleRule(start_type='percentile', start_value=0, start_color='4C9900',
                                                     mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                                     end_type='percentile', end_value=100, end_color='FF0000'))
        ws.conditional_formatting.add('A1:A1000',
                                      CellIsRule(operator='equal', formula=['MAX(A1:A1000)'], font=Font(bold=True)))
        wb.save(xlsx_out)

        counter = counter + 1
    else:
        wb = load_workbook(filename=xlsx_out)
        writer = pd.ExcelWriter(xlsx_out)
        writer.book = wb
        pd.read_csv(csv_out).to_excel(writer, sheet_name=junction, index=False)
        writer.save()
        writer.close()
        wb = load_workbook(filename=xlsx_out)
        ws = wb[junction]
        ws.conditional_formatting.add('A1:A1000',
                                      ColorScaleRule(start_type='percentile', start_value=0, start_color='4C9900',
                                                     mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                                     end_type='percentile', end_value=100, end_color='FF0000'))
        ws.conditional_formatting.add('A1:A1000',
                                      CellIsRule(operator='equal', formula=['MAX(A1:A1000)'], font=Font(bold=True)))
        wb.save(xlsx_out)

        counter = counter + 1

    print("\n***** Junction " + str(counter) + " of " + str(len(junction_list)) + " complete! *****\n")

# Copy final file(s) from 'temp_dir' to 'output_dir'; clear 'temp_dir'
shutilb.copy(csv_out_2, output_dir)

for filename in os.listdir(temp_dir):
    file_path = os.path.join(temp_dir, filename)
    try:
        if os.path.isfile(file_path) or os.path.islink(file_path):
            os.unlink(file_path)
        elif os.path.isdir(file_path):
            shutilb.rmtree(file_path)
    except Exception as e:
        print('Failed to delete %s from TempDir. Reason: %s' % (file_path, e))

print("****************************************************************")
print("Storm locations have been optimized for " + str(len(junction_list)) + " junctions.")
print("The script results have been saved in: " + output_dir)
print("Run Complete.")
print("****************************************************************")
